# -*- coding: utf-8 -*-
"""
Extraction des données réelles du chantier d'Odienné vers un fichier JSON
exploitable par le seeder Laravel.

Source : BASE DE CALCUL D'AVANCEMENT PHYSIQUE DU PROJET JUIN 2026.xlsx,
document annexe du rapport mensuel n° 14 du groupement TAEP/IETF.

  python scripts/extraire_odienne.py <classeur.xlsx> <sortie.json>
"""
import json
import re
import sys
import unicodedata
import warnings
from datetime import date, datetime

import openpyxl

warnings.filterwarnings('ignore')

MOIS = {
    'janv': 1, 'jan': 1, 'fev': 2, 'fév': 2, 'mars': 3, 'mar': 3, 'avr': 4,
    'mai': 5, 'juin': 6, 'juil': 7, 'aout': 8, 'août': 8, 'sept': 9,
    'oct': 10, 'nov': 11, 'dec': 12, 'déc': 12,
}


def sans_accent(t):
    return ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')


def cle(t):
    """Clé de rapprochement entre les deux feuilles, tolérante aux variantes."""
    t = str(t or '').replace('Œ', 'OE').replace('œ', 'oe')
    t = re.sub(r'[^A-Z]', '', sans_accent(t).upper())
    # Fautes de frappe du classeur d'origine.
    return (t.replace('HERBERGEMENT', 'HEBERGEMENT')
             .replace('ENSEINGNANTS', 'ENSEIGNANTS')
             .replace('REPLEMENT', 'REPLIEMENT'))


# La feuille de planning ne nomme pas toujours l'ouvrage comme la feuille
# d'avancement ; ces correspondances sont établies à la lecture des deux.
CORRESPONDANCES = {
    'PRESIDENCEADMINISTRATIONCENTRALECELODATCENTERPCSECURITE': ['BATIMENTPRESIDENCE'],
    'UFRSCOLARITECENTRALE': ['UFRSCOLARITECENTRALE'],
    # Les trois logements de fonction ne forment qu'un poste d'avancement :
    # on retient la première date de démarrage prévue et la dernière échéance.
    'LOGEMENTSDEFONCTIONVILLADUPRESIDENTVILLASDESDRETVICEPRESIDENTLOGEMENTDASTREINTE': [
        'LOGEMENTDUPRESIDENT', 'LOGEMENTSDUVICEPRETDIRECTEURS', 'LOGEMENTSDASTREINTE',
    ],
    'TERRASSEMENTS': ['TERRASSEMENTSPLATEFORME'],
    'VOIRIESAIRESDESTATIONNEMENTETPARKING': ['VOIRIESAIRESDESTATIONNEMENTET'],
    'ENVIRONNEMENTAMENAGEMENTPAYSAGER': ['ENVIRONNEMENTAMENAGEMENTPAYSAGER'],
}


def fusionner(calendriers):
    """Plusieurs tâches pour un même ouvrage : enveloppe la plus large."""
    debuts = [c['debut_prevu'] for c in calendriers if c['debut_prevu']]
    fins = [c['fin_prevue'] for c in calendriers if c['fin_prevue']]
    reels = [c['debut_reel'] for c in calendriers if c['debut_reel']]
    retards = [c['retard_mois_source'] for c in calendriers if c['retard_mois_source']]

    return {
        'duree_jours': max((c['duree_jours'] for c in calendriers if c['duree_jours']), default=None),
        'debut_prevu': min(debuts) if debuts else None,
        'fin_prevue': max(fins) if fins else None,
        'debut_reel': min(reels) if reels else None,
        'retard_mois_source': max(retards) if retards else None,
        'observations': calendriers[0]['observations'],
    }


def jour(v):
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return None


def mois_texte(v):
    """« Nov. 2025 » ou « Mars.. 2025 » → 2025-11-01, par convention le 1er."""
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    if not isinstance(v, str):
        return None
    m = re.match(r'\s*([A-Za-zéûôàè]+)\.*\s*(\d{4})', sans_accent(v))
    if not m:
        return None
    mois = MOIS.get(m.group(1).lower()[:5]) or MOIS.get(m.group(1).lower()[:4]) or MOIS.get(m.group(1).lower()[:3])
    return f'{int(m.group(2)):04d}-{mois:02d}-01' if mois else None


def entier(v):
    if isinstance(v, (int, float)):
        return int(round(v))
    if isinstance(v, str):
        m = re.search(r'\d+', v)
        return int(m.group()) if m else None
    return None


def pct(v):
    """Le classeur exprime les taux en fraction ; l'application les veut en %."""
    return round(v * 100, 2) if isinstance(v, (int, float)) else None


# ─────────────────────────────────────────────────── feuille « progress »

# Postes de premier niveau ; leur somme fait 100 % du marché.
GROUPES = {
    'INSTALLATIONSDECHANTIER', 'BATIMENTS', 'MOBILIERSETEQUIPEMENTS',
    'TERRASSEMENTSVRDAMENAGEMENTSEXTERIEURS', 'EFFICACITEENERGETIQUE',
}
# Postes sans poids, hors périmètre de la pondération.
IGNORES = {'ETUDESDEXECUTION', 'REPLIEMENTREMISEENETATDUCHANTIER'}
# Ces groupes sont suivis comme un seul ouvrage, sans descendre à leurs postes.
GROUPES_FEUILLES = {'INSTALLATIONSDECHANTIER', 'MOBILIERSETEQUIPEMENTS', 'EFFICACITEENERGETIQUE'}


def lire_progress(ws, arrete):
    """Ouvrages pondérés et séries mensuelles prévu/réel."""
    # Le classeur projette jusqu'à la fin du marché ; on s'arrête à la date
    # d'arrêté, au-delà de laquelle rien n'est constaté.
    periodes = []
    for v in next(ws.iter_rows(min_row=3, max_row=3, values_only=True)):
        if isinstance(v, (datetime, date)) and v.strftime('%Y-%m') <= arrete:
            periodes.append(v.strftime('%Y-%m'))
    # Colonnes des séries : première colonne mensuelle repérée sur la même ligne.
    ligne3 = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    col0 = next(i for i, v in enumerate(ligne3) if isinstance(v, (datetime, date)))

    lignes = list(ws.iter_rows(min_row=5, max_row=241, values_only=True))
    ouvrages, projet = [], {}
    groupe, poids_groupe = None, 100.0

    for i, r in enumerate(lignes):
        nom, poids = r[3], r[8]
        if not nom or not isinstance(poids, (int, float)):
            continue
        k = cle(nom)
        if k in IGNORES or 'AVANCEMENTPHYSIQUEGLOBAL' in k:
            if 'AVANCEMENTPHYSIQUEGLOBAL' in k and not projet:
                projet = series(lignes, i, col0, periodes)
            continue

        if k in GROUPES:
            groupe, poids_groupe = k, poids * 100
            if k not in GROUPES_FEUILLES:
                continue  # on descendra à ses postes
            groupe = None  # ses postes ne sont pas suivis séparément
        elif groupe is None:
            continue  # poste d'un groupe suivi globalement

        # Poids ramené au marché entier : celui du groupe × celui du poste.
        poids_projet = poids * 100 if k in GROUPES_FEUILLES else poids_groupe * poids
        ouvrages.append({
            'nom': re.sub(r'\s+', ' ', str(nom)).strip(),
            'cle': k,
            'groupe': groupe,
            'poids': round(poids_projet, 4),
            'debut_planning': jour(r[4]),
            'fin_planning': jour(r[5]),
            **series(lignes, i, col0, periodes),
        })

    return periodes, projet, ouvrages


def series(lignes, i, col0, periodes):
    """Les quatre lignes qui suivent un poste portent ses séries mensuelles."""
    out = {'prevu': {}, 'reel': {}}
    for r in lignes[i + 1:i + 7]:
        libelle = sans_accent(str(r[3] or '')).lower()
        cible = 'prevu' if libelle.startswith('prevision cumul') else (
            'reel' if libelle.startswith('realisation cumul') else None)
        if not cible:
            continue
        for j, p in enumerate(periodes):
            v = pct(r[col0 + j]) if col0 + j < len(r) else None
            if v is not None:
                out[cible][p] = min(100.0, v)
    return out


# ────────────────────────────────────────────────── feuille « récap mois »

def lire_calendrier(ws):
    """Calendrier contractuel par ouvrage : durée, dates prévues, début réel."""
    cal = {}
    for r in ws.iter_rows(min_row=18, max_row=60, max_col=11, values_only=True):
        nom = r[2]
        if not nom or not isinstance(r[3], (int, float)):
            continue
        cal[cle(nom)] = {
            'duree_jours': entier(r[3]),
            'debut_prevu': jour(r[4]),
            'fin_prevue': jour(r[5]),
            'debut_reel': mois_texte(r[6]),
            'retard_mois_source': entier(r[7]),
            'observations': re.sub(r'\s+', ' ', str(r[10] or '')).strip() or None,
        }
    return cal


# ──────────────────────── feuille « détail facturation PFO tache »

CORRESPONDANCES_FIN = {
    'EAUXUSEES': 'ASSAINISSEMENENTEAUXUSEESSTEP',
    'ADDUCTIONENEAUXPOTABLEAEP': 'AEPHORSCHATEAUDEAU',
    'EAUXPLUVIALE': 'EAUPLUVIALE',
    'RESEAUFIBRE': 'RESEAUTELECOMFIBREOPTIQUE',
    'SIGNALISATIONINTERIEURE': 'SIGNALETIQUEINTERIEURE',
    'ENVIRONNEMENTAIREDEJEUX': 'AIREDEJEUX',
    'ENVIRONNEMENTAMENAGEMENTPAYSAGER': 'AMENAGEMENTPAYSAGER',
    'MOBILIERSETEQUIPEMENTS': 'TOTALMOBILIEREQUIPEMENT',
    'EFFICACITEENERGETIQUE': 'SOUSTOTALEFFICACITEENERGETIQUE',
}
TOTAUX = ('SOUSTOTAL', 'TOTALGENERAL', 'TOTALVRD')


def lire_revision(ws):
    """Provision pour revision de prix : facturee, rattachee a aucun ouvrage."""
    for r in ws.iter_rows(min_row=14, max_row=70, max_col=7, values_only=True):
        if cle(r[2]) == 'PROVISIONPOURREVISIONDEPRIX' and isinstance(r[5], (int, float)):
            return round(float(r[5]), 2)
    return 0.0


def lire_enveloppes(ws):
    """Montant contractuel et facturation cumulée de chaque ouvrage."""
    env = {}
    for r in ws.iter_rows(min_row=14, max_row=70, max_col=7, values_only=True):
        nom = r[2]
        montant = float(r[3]) if isinstance(r[3], (int, float)) else 0.0
        facture = float(r[5]) if isinstance(r[5], (int, float)) else 0.0
        # Un poste peut etre facture sans montant contractuel propre.
        if not isinstance(nom, str) or not nom.strip() or (montant <= 0 and facture <= 0):
            continue
        k = cle(nom)
        if any(k.startswith(t) for t in TOTAUX) and k not in CORRESPONDANCES_FIN.values():
            continue
        if k in env:
            continue
        env[k] = {'montant': round(montant, 2), 'facture': round(facture, 2)}
    return env


def enveloppe(k, env):
    if k in CORRESPONDANCES_FIN:
        return env.get(CORRESPONDANCES_FIN[k])
    if k in env:
        return env[k]
    if len(k) < 7:
        return None
    for autre, v in env.items():
        if len(autre) >= 7 and (autre.startswith(k[:7]) or k.startswith(autre[:7])):
            return v
    return None


# ─────────────────────────────────────────────── feuille « facturation PFO »

def lire_decomptes(ws):
    """Décomptes de l'entreprise, avance de démarrage comprise."""
    out = []
    for r in ws.iter_rows(min_row=8, max_row=30, max_col=12, values_only=True):
        num, periode, nature, brut = r[2], r[3], r[4], r[5]
        if not isinstance(brut, (int, float)) or not isinstance(periode, (datetime, date)):
            continue
        recup = sum(float(v) for v in (r[6], r[7]) if isinstance(v, (int, float)))
        out.append({
            'numero': str(entier(num)) if entier(num) is not None else 'ACOMPTE',
            'date': jour(periode),
            'nature': re.sub(r'\s+', ' ', str(nature or 'Travaux')).strip(),
            'brut': round(float(brut), 2),
            'recuperation_avance': round(recup, 2),
            'net_paye': round(float(r[8]), 2) if isinstance(r[8], (int, float)) else None,
            'regle': str(r[9] or '').strip().lower().startswith('pay'),
        })
    return out


def main():
    source, sortie = sys.argv[1], sys.argv[2]
    arrete = '2026-06'
    w = openpyxl.load_workbook(source, data_only=True)

    periodes, projet, ouvrages = lire_progress(w['progress'], arrete)
    calendrier = lire_calendrier(w['récap mois'])
    enveloppes = lire_enveloppes(w['détail facturation PFO tache'])
    decomptes = lire_decomptes(w['facturation PFO'])

    manquants = []
    sans_enveloppe = []
    for o in ouvrages:
        k = o.pop('cle')
        trouves = [calendrier[c] for c in CORRESPONDANCES.get(k, [k]) if c in calendrier]
        if not trouves:
            trouves = [v for c, v in calendrier.items()
                       if c.startswith(k[:12]) or k.startswith(c[:12])]
        if trouves:
            o.update(fusionner(trouves))
        else:
            manquants.append(o['nom'])

        env = enveloppe(k, enveloppes)
        o['enveloppe'] = env['montant'] if env else None
        o['facture_cumulee'] = env['facture'] if env else None
        if env is None:
            sans_enveloppe.append(o['nom'])

    total = round(sum(o['poids'] for o in ouvrages), 2)
    data = {
        'source': 'Base de calcul d’avancement physique — juin 2026, groupement TAEP/IETF',
        'arrete_au': '2026-06-30',
        'periodes': periodes,
        'projet': projet,
        'ouvrages': ouvrages,
        'decomptes': decomptes,
        'ponderation_totale': total,
        'revision_facturee': lire_revision(w['détail facturation PFO tache']),
        'anomalies': [],
    }
    with open(sortie, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f'{len(ouvrages)} ouvrages, somme des pondérations = {total} %')
    print(f'{len(periodes)} périodes de {periodes[0]} à {periodes[-1]}')
    print(f'{len(decomptes)} décomptes')
    print(f'Projet : prévu {projet["prevu"].get(periodes[-1])} % / réel {projet["reel"].get(periodes[-1])} %')
    print(f'Enveloppes : {sum(1 for o in ouvrages if o["enveloppe"])} / {len(ouvrages)} ouvrages')
    print(f'Facturé cumulé : {sum(o["facture_cumulee"] or 0 for o in ouvrages):,.0f}')
    if manquants:
        print('Sans calendrier : ' + ', '.join(manquants))
    if sans_enveloppe:
        print('Sans enveloppe : ' + ', '.join(sans_enveloppe))


if __name__ == '__main__':
    main()
